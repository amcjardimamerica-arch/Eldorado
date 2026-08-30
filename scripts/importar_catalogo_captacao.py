"""Converte a planilha do acervo em catálogos auditáveis por nível e tipo.

Uso manual: python scripts/importar_catalogo_captacao.py caminho/planilha.xlsx
Os registros do acervo são pistas, não editais confirmados.
"""
from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]


def clean(value) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def fold(value: str) -> str:
    return unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower()


def levels(area: str, source: str, place: str, kind: str) -> list[str]:
    text = fold(" ".join((area, source, place, kind)))
    if "internacional" in text or "grant internacional" in text:
        return ["internacional"]
    found = []
    rules = {
        "municipal": ("municipal", "goiania", "prefeitura", "camara municipal", "fmdca", "cmdca"),
        "estadual": ("estadual", "goias", "alego", "secult goias", "seds", "governo de goias"),
        "federal": ("federal", "ministerio", "transferegov", "conanda", "bndes", "mjps", "mjsp", "mma", "receita federal"),
        "privada": ("privad", "empresa", "fundacao", "instituto", "plataforma", "crowdfunding", "cooperativa", "doacao recorrente"),
    }
    for level, words in rules.items():
        if any(word in text for word in words):
            found.append(level)
    return found or ["a_validar"]


def subtype(kind: str, area: str) -> str:
    text = fold(f"{kind} {area}")
    if "internacional" in text: return "grants_internacionais"
    if "emenda" in text: return "emendas"
    if "mrosc" in text or "chamamento" in text or "edital" in text: return "editais_chamamentos"
    if "incentivo fiscal" in text: return "incentivos_fiscais"
    if "fundo" in text: return "fundos"
    if "tac" in text or "judicial" in text or "pena" in text: return "justica_destinacoes"
    if "investimento social" in text or "doacao" in text or "patrocin" in text: return "doacoes_patrocinios"
    if "plataforma" in text: return "plataformas_radar"
    return "outros"


def main(xlsx: Path) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Instale openpyxl apenas para esta importação.") from exc
    sheet = load_workbook(xlsx, read_only=True, data_only=True).active
    entries = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        number, area, source, place, amount, kind = (clean(v) for v in row[:6])
        if not source:
            continue
        item_levels = levels(area or "", source, place or "", kind or "")
        entries.append({
            "id_acervo": f"captacao-{int(float(number)):03d}" if number else f"captacao-{len(entries)+1:03d}",
            "area_original": area,
            "fonte_programa": source,
            "onde_captar_original": place,
            "valor_texto_nao_verificado": amount,
            "tipo_original": kind,
            "niveis_inferidos": item_levels,
            "tipo_catalogo": subtype(kind or "", area or ""),
            "status": "pista_do_acervo_pendente_url_primaria",
            "origem": "06 - Captação Fazer.rar/Planilha projetos ideias.xlsx",
        })
    base = {"versao": 1, "importado_em": date.today().isoformat(), "total": len(entries), "aviso": "Dados do anexo são pistas. Valores, elegibilidade e disponibilidade exigem fonte primária atual.", "itens": entries}
    root = ROOT / "catalogo_captacao"
    root.mkdir(parents=True, exist_ok=True)
    (root / "catalogo_anexo.json").write_text(json.dumps(base, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    grouped = defaultdict(list)
    for item in entries:
        for level in item["niveis_inferidos"]:
            grouped[(level, item["tipo_catalogo"])].append(item)
    for (level, kind), items in grouped.items():
        folder = root / level / kind
        folder.mkdir(parents=True, exist_ok=True)
        payload = {"nivel": level, "tipo": kind, "total": len(items), "itens": items}
        (folder / "catalogo.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"total": len(entries), "pastas": len(grouped)}, ensure_ascii=False))


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("uso: importar_catalogo_captacao.py planilha.xlsx")
    main(Path(sys.argv[1]))
